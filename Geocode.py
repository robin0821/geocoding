# -*- coding: utf-8 -*-
"""
Created on Thu Apr 13 20:49:24 2017

@author: Administrator
"""

from openpyxl import load_workbook
from googleplaces import GooglePlaces, types
import time
import numpy as np

my_api_key = 'AIzaSyAaXc0c1EB2LGIXtokrmaFVTqherSX_w2I'
wz_api_key = 'AIzaSyA4ctAYpLK--cuSES1UE-GhZxzpk3IHMgI'
google_places = GooglePlaces(wz_api_key)

wb = load_workbook(filename = 'sub_rank_Clean_V01.xlsx')
ws = wb.get_sheet_by_name('sub_rank')

#rows = ws.rows
#columns = ws.columns
#print(rows,columns)

for row in range(297, 2237):
    school = ws['C' + str(row)].value + ' Melbourne, Australia'
    
    query_result = google_places.text_search(school, types=[types.TYPE_SCHOOL])
    
    i = 1
    for place in query_result.places:
        place.get_details()
    #    print(place.name)
        if i >= 2:
            break
        else:
            addr = place.formatted_address
            website = place.website
            telephone = place.international_phone_number
            cord = place.geo_location
            latitude = cord['lat']
            longitude = cord['lng']
            ws['J' + str(row)].value = addr
            ws['K' + str(row)].value = website
            ws['L' + str(row)].value = telephone
            ws['M' + str(row)].value = latitude
            ws['N' + str(row)].value = longitude
        wb.save('sub_rank_Clean_v01.xlsx')
        print(school + " is finished!")
        i += 1
        time.sleep(np.random.rand()*20)

