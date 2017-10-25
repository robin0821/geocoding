# -*- coding: utf-8 -*-
"""
Created on Sun Apr 16 09:53:50 2017

@author: Administrator
"""

from openpyxl import load_workbook
from geopy.geocoders import Nominatim
import time

wb = load_workbook(filename = 'api1.xlsx')
ws = wb.get_sheet_by_name('API')
geolocator = Nominatim(scheme='http')

for row in range(438, 439):
    address = ws['G' + str(row)].value
    try:
        location = geolocator.geocode(address)
        try:
            lat = location.latitude
        except:
            lat = "n.a."
        try:
            lon = location.longitude
        except:
            lon = "n.a."
        ws['E' + str(row)].value = lat
        ws['F' + str(row)].value = lon
        wb.save('api1.xlsx')
        print(address)
        print("row " + str(row) + str(lat) + str(lon) + " is written!")
    except:
        print(address)
        print(str(row) + " is skipped!")
        continue

    #print(lat,lon)
    

    time.sleep(1.5)
    